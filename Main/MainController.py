from Screens import mainscreen
from Models.Dictionaries import SKILLS
from Models import Product
from Functions import SetCriticalOperation, WorkerAssigner
from Functions.ExcelDataLoader import ExcelDataLoader
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime


class MainController:
    def __init__(self):
        self.__screenController = None
        self.__products = []  # Holds product objects
        self.__jigs = []  # Holds jig objects
        self.__workers = []  # Holds worker objects
        self.__all_critical_operations = []  # [(product , crital operations)]
        self.__dataLoaderObject = ExcelDataLoader()  # Creates DataLoaderObject
        self.__dataLoaderObject.set_products(self.__products)
        self.__dataLoaderObject.set_jigs(self.__jigs)
        self.__dataLoaderObject.set_workers(self.__workers)
        self.__ScheduleObject = WorkerAssigner.Schedule()
        self.__critical_op_check_list = []
        self.__assigned_operations = {}
        self.__operation_workers = {}

    def is_operation_assigned_to_interval(self, operation_name, time_interval, product=None):
        """Checks if an operation is already assigned to a time interval, optionally for a specific product"""
        # Check the current assignments in the time interval
        for assignment in time_interval.get_assignments():
            if len(assignment) >= 3:  # Make sure assignment has enough elements
                assigned_jig, assigned_product, assigned_operation, assigned_workers = assignment
                if assigned_operation.get_name() == operation_name:
                    # If product is specified, check if it matches
                    if product is None or assigned_product == product:
                        return True
        return False

    def add_operation_assignment(self, operation_name, time_interval, workers_count):
        """Tracks that an operation has been assigned to a time interval"""
        if operation_name not in self.__assigned_operations:
            self.__assigned_operations[operation_name] = []

        self.__assigned_operations[operation_name].append((time_interval, workers_count))

    def clear_assignments_tracking(self):
        """Clears the operation assignments tracking (call before starting a new scheduling)"""
        self.__assigned_operations = {}

    def create_product(self, serialNumber=None):
        Product.create_product(self.__products, serialNumber)

    def delete_product(self, serialNumber):
        for product in self.__products:
            if product.get_serial_number() == serialNumber:
                product.get_current_jig().set_state(None)
                product.get_current_jig().set_assigned_product(None)
                self.__products.remove(product)

    def get_product_list(self):
        return self.__products

    def get_product(self, serialNumber):
        for product in self.__products:
            if product.get_serial_number() == serialNumber:
                return product

    def get_jig(self, name):
        for jig in self.__jigs:
            if jig.get_name() == name:
                return jig

    def get_jigs(self):
        return self.__jigs

    def get_jig(self, _name):
        for jig in self.__jigs:
            if jig.get_name() == _name:
                return jig

    def get_workers(self):
        return self.__workers

    def get_worker(self, _reg_no):
        for worker in self.__workers:
            if worker.get_registration_number() == _reg_no:
                return worker

    def get_data_loader_object(self):
        return self.__dataLoaderObject

    def get_ScheduleObject(self):
        return self.__ScheduleObject

    def calculate_required_worker(self, sn):
        product = self.get_product(sn)
        for operation in product.get_operations():
            required_man = operation.get_min_workers()
            while required_man <= operation.get_max_workers():
                if (operation.get_required_man_hours()/required_man) <= 7.5:
                    operation.set_required_worker(required_man)
                    break
                else:
                    required_man = required_man + 1
            print(f"Op {operation.get_name()} req shift: {operation.get_required_worker()}")

    def calculate_operating_duration(self, sn):
        product = self.get_product(sn)
        for operation in product.get_operations():
            man_hours = operation.get_required_man_hours()
            workers = operation.get_required_worker()

            # Bir gün için 7.5 saat çalışma
            duration = man_hours / (7.5 * workers)

            # Hesaplamaları yazdır
            print(f"Operation {operation.get_name()} calculation:")
            print(f"  Required man-hours: {man_hours}")
            print(f"  Required workers: {workers}")
            print(f"  Duration formula: {man_hours} / (7.5 * {workers}) = {duration}")

            operation.set_operating_duration(duration)
            operation.set_remaining_duration(duration)  # Initialize remaining duration

        self.print_operation_durations()

    def print_operation_durations(self):
        for product in self.__products:
            for operation in product.get_operations():
                print(f"Operation {operation.get_name()} duration: {operation.get_operating_duration()}")

    def set_all_previous_operations(self, sn):
        product = self.get_product(sn)
        all_ops = product.get_operations()
        for op in product.get_operations():
            self.find_all_previous_operations(op, all_ops)

    def find_all_previous_operations(self, operation, all_operations, visited=None):
        if visited is None:
            visited = set()

        if operation.get_name() in visited:
            return

        visited.add(operation.get_name())
        prev_list = operation.get_previous_operations()

        for pred in operation.get_predecessors():
            if pred not in prev_list:
                prev_list.append(pred)
            self.find_all_previous_operations(pred, all_operations, visited)
            for pred_prev in pred.get_previous_operations():
                if pred_prev not in prev_list:
                    prev_list.append(pred_prev)

        operation.set_previous_operations(prev_list)

    def calculate_product_progress(self, serial_number):
        product = self.get_product(serial_number)
        total_duration = 0
        for operation in product.get_operations():
            total_duration = total_duration + operation.get_required_man_hours()
        applied_duration = 0
        for operation in product.get_operations():
            if operation.get_completed():
                applied_duration = applied_duration + operation.get_required_man_hours()
        progress = applied_duration/total_duration*100
        product.set_progress(progress)

        print(f"product {product.get_serial_number()} progress % : {product.get_progress()}")

    def remove_completed_predecessors(self, _sn):
        product = self.get_product(_sn)

        # Önce tüm tamamlanmış operasyonları bul
        completed_operations = []
        for operation in product.get_operations():
            if operation.get_completed():
                completed_operations.append(operation)

        # Her operasyon için tamamlanmış öncülleri güncelle
        for operation in product.get_operations():
            if not operation.get_completed():  # Sadece tamamlanmamış operasyonları güncelle
                # Öncül listelerini klonla
                uncompleted_predecessors = []

                # Tamamlanmamış öncülleri tespit et
                for pred in operation.get_predecessors():
                    if not pred.get_completed():
                        uncompleted_predecessors.append(pred)

                # Tamamlanmamış öncülleri güncelle
                operation.set_uncompleted_predecessors(uncompleted_predecessors)

                # Eğer artık tamamlanmamış öncül kalmadıysa, işaretleyelim
                if len(uncompleted_predecessors) == 0:
                    print(f"Operation {operation.get_name()} has no more incomplete predecessors")

    # CPM calculation
    def set_critical_operations(self, _sn):
        """
        Calculates the critical path for a product using remaining operation durations.
        Only considers operations that are not yet completed.

        Args:
            _sn: Product serial number
        """
        tasks = []
        product = self.get_product(_sn)
        calculator = SetCriticalOperation.Graph()

        # Sadece tamamlanmamış operasyonları CPM hesaplamasına dahil et
        for operation in product.get_operations():
            if not operation.get_completed():
                task = operation.get_name()

                # Operasyon süresini kontrol et, kalan süreyi kullan
                # Eğer kalan süre tanımlanmamışsa veya sıfırsa, tam süreyi kullan
                if operation.get_remaining_duration() is None or operation.get_remaining_duration() <= 0:
                    # Eğer kalan süre tanımlanmamışsa, tam süreyi ata
                    operation.set_remaining_duration(operation.get_operating_duration())

                # En az 0.01 süre olmasını sağla (CPM hesaplamasında sıfır süre problem çıkarabilir)
                duration = max(operation.get_remaining_duration(), 0.01)

                # Sadece tamamlanmamış öncülleri ekle
                dependencies = []
                for op in operation.get_uncompleted_predecessors():
                    dependencies.append(op.get_name())

                # Grafiğe ekle
                calculator.add_task(task, duration, dependencies)
                tasks.append(task)

        # Kritik operasyonları hesapla
        critical_operations, earliest_start, latest_finish = calculator.find_critical_operations()

        print("Kritik Operasyonlar:", critical_operations)
        print("En Erken Başlama Zamanları:", earliest_start)
        print("En Geç Tamamlanma Zamanları:", latest_finish)

        # Kritik operasyonları listeye ekle
        critical_op_obj_list = []
        for op_name in critical_operations:
            op_obj = product.get_operation_by_name(op_name)
            if op_obj and not op_obj.get_completed():  # Tamamlanmamış olduğundan emin ol
                op_obj.set_early_start(earliest_start[op_name])
                op_obj.set_late_finish(latest_finish[op_name])
                if op_obj.get_early_start() == 0:
                    critical_op_obj_list.append(op_obj)

        # Kritik operasyonları ürüne ata
        product.append_critical_operations(critical_op_obj_list)

    def sort_operations_by_duration(self):
        for product in self.__products:
            criticalops = product.get_critical_operations()
            sorted_criticalops = sorted(criticalops, key=lambda op: op.get_operating_duration())
            product.append_critical_operations(sorted_criticalops)

    def sort_products_by_progress(self):
        sorted_products = sorted(self.__products, key=lambda product: product.get_progress() or 0, reverse=True)
        self.__products = sorted_products

    def get_all_critical_operations(self):
        critical_ops = []
        seen_operations = set()  # Instead of using just operation name, use (product_serial, op_name) tuples
        product_op_count = {}  # Her ürün için atanan operasyon sayısını takip etmek için

        # Önce her ürün için kritik operasyonları topla
        product_critical_ops = {}

        for product in self.__products:
            product_sn = product.get_serial_number()
            product_op_count[product_sn] = 0
            product_critical_ops[product_sn] = []

            for op in product.get_critical_operations():
                # Kalan süresi olan operasyonları kontrol et
                has_remaining_time = op.get_remaining_duration() is not None and op.get_remaining_duration() > 0.001

                # Operasyon tamamlanmadı VEYA kalan süresi varsa ve öncülleri tamamlandıysa
                if (not op.get_completed() or has_remaining_time) and len(op.get_uncompleted_predecessors()) == 0:
                    # Eğer tamamlandı olarak işaretlenmişse ama kalan süresi varsa, durumu düzelt
                    if op.get_completed() and has_remaining_time:
                        print(
                            f"Fixing inconsistency: Operation {op.get_name()} marked as completed but has remaining time: {op.get_remaining_duration()}")
                        op.set_completed(False)  # Completed flag'i düzelt

                    # Create a unique identifier for each operation based on product and operation name
                    operation_key = (product_sn, op.get_name())

                    # Aynı ürün+operasyon çifti daha önce eklenmiş mi kontrol et
                    if operation_key not in seen_operations:
                        seen_operations.add(operation_key)
                        product_critical_ops[product_sn].append((product, op))
                        print(
                            f"Critical operation: {op.get_name()} for product {product_sn}, remaining: {op.get_remaining_duration()}")
                    else:
                        print(f"Skipping duplicate critical operation: {op.get_name()} for product {product_sn}")

        # Şimdi ürünler arasında dönüşümlü olarak operasyon seç
        # Önce ürünleri progress'e göre sırala (progress'i düşük olanlar önce)
        products_by_progress = sorted(self.__products, key=lambda p: p.get_progress() or 0)

        # Her üründen sırayla birer operasyon ekleyerek ilerleme sağla
        while any(product_critical_ops.values()):
            for product in products_by_progress:
                product_sn = product.get_serial_number()
                if product_critical_ops[product_sn]:
                    # Bu üründen bir operasyon al ve ana listeye ekle
                    critical_ops.append(product_critical_ops[product_sn].pop(0))
                    product_op_count[product_sn] += 1

                    # Eğer bir üründen yeterince operasyon eklediyse ara
                    if product_op_count[product_sn] >= 2 and any(len(ops) > 0 for ops in product_critical_ops.values()):
                        break

        print("Critical operations for scheduling, balanced across products:")
        for product, op in critical_ops:
            print(
                f"Product: {product.get_serial_number()}, Operation: {op.get_name()}, Remaining: {op.get_remaining_duration()}")

        return critical_ops

    def set_schedule_attributes(self):
        self.__ScheduleObject.set_start_date(self.screenController.get_schedule_start())
        self.__ScheduleObject.set_end_date(self.screenController.get_schedule_end())
        self.__ScheduleObject.set_start_shift(self.screenController.get_starting_shift())
        self.__ScheduleObject.set_working_order(self.screenController.get_working_order_value())
        self.__ScheduleObject.create_time_intervals()
        self.assign_workers_to_time_intervals()

    def get_previous_workers_for_operation(self, operation_name):
        """
        Bir operasyona daha önce atanmış işçileri döndürür
        """
        if operation_name in self.__operation_workers:
            return self.__operation_workers[operation_name]
        return []

    def update_operation_workers(self, operation_name, workers):
        """
        Bir operasyon için atanan işçileri günceller
        """
        self.__operation_workers[operation_name] = workers

    def clear_operation_workers(self):
        """
        İşçi takibini temizler (yeni bir çizelgeleme başlamadan önce çağrılmalı)
        """
        self.__operation_workers = {}
    def assign_workers_to_time_intervals(self):
        if not self.__ScheduleObject or not self.__workers:
            raise ValueError("Schedule and workers must be set.")

        for date_obj in self.__ScheduleObject.dates:
            for time_interval in date_obj.time_intervals:
                # TimeInterval'ın tarih, vardiya ve zaman aralığını al
                interval_date = date_obj.get_date()  # Zaman aralığının tarihi
                interval_shift = time_interval.shift  # Vardiya
                interval_start_time = time_interval.interval[0]  # Zaman aralığının başlangıcı
                interval_end_time = time_interval.interval[1]  # Zaman aralığının bitişi

                # Uygun çalışanları bul
                available_workers = []
                for worker in self.__workers:
                    # Çalışanın off-day'lerini kontrol et
                    off_days = worker.get_off_days()
                    if off_days:
                        off_start_date = datetime.strptime(off_days[0], "%d.%m.%Y").date()
                        off_end_date = datetime.strptime(off_days[1], "%d.%m.%Y").date()
                        if off_start_date <= interval_date <= off_end_date:
                            continue  # Çalışan bu tarihte off-day'de, atama yapma

                    for schedule_entry in worker.get_shift_schedule():
                        schedule_date, schedule_shift, available_hours = schedule_entry

                        # Tarih ve vardiya eşleşiyor mu kontrol et
                        if schedule_date == interval_date and schedule_shift == interval_shift:
                            # Zaman aralığı da uyuyor mu kontrol et
                            for hours in available_hours:
                                if hours[0] <= interval_start_time and hours[1] >= interval_end_time:
                                    available_workers.append(worker)
                                    break

                # TimeInterval'ın available_workers listesini güncelle
                time_interval.available_workers = available_workers

    def previous_operation_control(self, operation, time_interval, product=None):
        """
        Checks if an operation can be scheduled in the given time interval.
        Verifies that all predecessors are completed and not running in the same interval.
        Prevents assigning the same operation multiple times in the same interval.

        Args:
            operation: The operation to be scheduled
            time_interval: The time interval to check
            product: The product this operation belongs to (added parameter)

        Returns:
            True if the operation can be scheduled, False otherwise
        """
        # If product wasn't passed, get it from the operation's context
        # This maintains compatibility with existing code
        if product is None:
            # Find the product that contains this operation
            for p in self.__products:
                if operation in p.get_operations():
                    product = p
                    break

        if product is None:
            print(f"Warning: Could not determine product for operation {operation.get_name()}")
            return False

        # Check if this operation is already assigned to this interval for this specific product
        if self.is_operation_assigned_to_interval(operation.get_name(), time_interval, product):
            print(
                f"Operation {operation.get_name()} for product {product.get_serial_number()} already assigned to this interval in previous_operation_control")
            return False

        # Check existing assignments in this interval
        assignments = time_interval.get_assignments()
        for assignment in assignments:
            if len(assignment) >= 3:  # Make sure assignment has enough elements
                assigned_jig, assigned_product, assigned_operation, assigned_workers = assignment

                # Only check assignments for the SAME product
                if assigned_product == product:
                    # Check if the same operation is already assigned in this interval for this product
                    if assigned_operation.get_name() == operation.get_name():
                        print(
                            f"Operation {operation.get_name()} for product {product.get_serial_number()} already has an existing assignment in this interval")
                        return False

                    # Check if any predecessor is running in this interval for this product
                    for prev_op in operation.get_previous_operations():
                        if assigned_operation.get_name() == prev_op.get_name():
                            print(
                                f"Predecessor {prev_op.get_name()} running in this interval for product {product.get_serial_number()}")
                            return False  # Predecessor is running in this interval

        # First check if all predecessors are completed for THIS product
        for pred in operation.get_predecessors():
            # Make sure the predecessor belongs to the same product
            if pred in product.get_operations() and not pred.get_completed():
                print(
                    f"Predecessor {pred.get_name()} not completed for operation {operation.get_name()} of product {product.get_serial_number()}")
                return False  # Cannot schedule if any predecessor is not completed

        return True  # All checks passed, operation can be scheduled

    def initiate_assignment(self, critical_op_list):
        """
        Initiates the assignment process for critical operations.
        Handles deadlock detection and supports partial assignment of operations.

        Args:
            critical_op_list: List of tuples (product, operation) to be scheduled
        """
        # Compare to previous operation list - if the same, we're stuck
        if critical_op_list == self.__critical_op_check_list and critical_op_list:
            print("Same operation list detected - forcing resolution")

            # Try a force assignment for the first operation in the list
            product, operation = critical_op_list[0]
            print(f"Forcing assignment for operation {operation.get_name()} of product {product.get_serial_number()}")

            # Get earliest available time interval
            intervals_list = self.get_ScheduleObject().get_sorted_time_intervals()
            if intervals_list:
                # Force assign to earliest available interval with any available workers
                for interval in intervals_list:
                    # Check if the operation is already assigned to this interval for this product
                    if self.is_operation_assigned_to_interval(operation.get_name(), interval, product):
                        print(f"Operation {operation.get_name()} already assigned to this interval - skipping")
                        continue

                    # Select best workers for assignment
                    selected_workers = self.select_best_workers_for_assignment(operation, interval,
                                                                               prefer_previous=True)

                    if selected_workers and len(selected_workers) >= operation.get_required_worker():
                        # Force the assignment regardless of other constraints
                        jig = product.get_current_jig()

                        # If jig is incompatible, try to change it
                        if not self.jig_compatibility_control(product, operation):
                            self.change_jig(product, operation)
                            jig = product.get_current_jig()

                        # Create the assignment and make sure operation is marked as completed
                        if self.create_assignment(interval, jig, product, operation, selected_workers):
                            # Update dependencies
                            product_sn = product.get_serial_number()
                            self.remove_completed_predecessors(product_sn)

                            print(f"Forced assignment successful for operation {operation.get_name()}")

                            # Reset critical ops check list to break comparison cycle
                            self.__critical_op_check_list = []
                            # Continue with the next operation
                            self.make_assignment_preparetions()
                            return

                # If no assignment could be made for first operation, remove it and continue
                print(f"Could not force assignment for operation {operation.get_name()}, removing from critical path")
                # Manually remove this operation from consideration
                operation.set_completed(True)  # Mark as completed to remove from critical path
                operation.set_remaining_duration(0)

                # Reset critical ops check list to break comparison cycle
                self.__critical_op_check_list = []
                self.make_assignment_preparetions()
                return

        # Store current operation list for comparison in next iteration
        self.__critical_op_check_list = critical_op_list

        # Process each critical operation
        for product, operation in critical_op_list:
            intervals_list = self.get_ScheduleObject().get_sorted_time_intervals()

            # 1. Find the latest finish time of predecessors (within the same product)
            latest_finish_time = self.find_latest_finish_time_of_predecessors(operation, product)

            # If latest_finish_time is None, it means some predecessors haven't been scheduled yet
            if latest_finish_time is None:
                print(
                    f"Skipping operation {operation.get_name()} for product {product.get_serial_number()} - predecessors not yet scheduled")
                continue  # Skip this operation for now

            # 2. Filter intervals that start after the latest finish time
            filtered_intervals = self.filter_intervals_after_time(intervals_list, latest_finish_time)

            if not filtered_intervals:
                print(
                    f"No valid intervals for operation {operation.get_name()} after predecessor completion time {latest_finish_time}")
                continue  # No valid intervals available, skip this operation

            # Initialize variables for tracking assignments
            # Use remaining_duration instead of operating_duration
            remaining_duration = operation.get_remaining_duration()
            if remaining_duration is None or remaining_duration <= 0:
                # If remaining_duration is not set or is 0, use the full operating_duration
                remaining_duration = operation.get_operating_duration()
                operation.set_remaining_duration(remaining_duration)  # Make sure remaining_duration is set

            print(
                f"Operation {operation.get_name()} for product {product.get_serial_number()} has remaining duration: {remaining_duration}")

            successfully_assigned = False

            # Operasyonun başlangıç tarihini takip etmek için
            operation_start_date = None
            max_days_to_consider = 2  # Başlangıçtan sonra en fazla kaç gün içinde tamamlanmalı

            # Check scheduling options: first try with same product, then without same product
            for same_product_mode in [True, False]:
                if successfully_assigned:
                    break

                for interval in filtered_intervals:
                    # Eğer operasyon başladıysa ve bu interval başlangıç tarihinden belirli gün sayısından fazla ilerdeyse atla
                    if operation_start_date is not None:
                        days_difference = (interval.get_date() - operation_start_date).days
                        if days_difference > max_days_to_consider:
                            continue

                    # Check if this operation is already assigned to this interval for this product
                    if self.is_operation_assigned_to_interval(operation.get_name(), interval, product):
                        print(
                            f"Operation {operation.get_name()} already assigned to interval {interval.get_date()} {interval.interval[0]}-{interval.interval[1]}")
                        continue  # Skip this interval

                    # Skip if a previous operation is still running in this interval
                    # Pass the product to ensure proper context for predecessor checks
                    if not self.previous_operation_control(operation, interval, product):
                        continue

                    # Skip this interval if we're in same_product_mode but no same product in interval,
                    # or if we're not in same_product_mode but there is same product in interval
                    has_same_product = self.same_product_control(product, interval)
                    if same_product_mode != has_same_product:
                        continue

                    # If same product is already in interval, check jig capacity
                    if has_same_product and not self.check_jig_capacity(product, operation, interval):
                        continue

                    # Verify we have enough workers
                    if not self.compatible_worker_number_check(operation, interval):
                        continue

                    # Handle jig compatibility
                    current_jig = product.get_current_jig()
                    if not has_same_product:  # Only change jig if not already using one for same product
                        if not self.jig_compatibility_control(product, operation):
                            self.change_jig(product, operation)

                    # Verify we can schedule the complete duration
                    jig = product.get_current_jig()
                    assignment_intervals = []
                    current_interval = interval
                    assignment_duration = 0

                    # Store worker assignments for each interval
                    interval_worker_assignments = {}
                    current_shift = current_interval.get_shift()

                    # Operasyon başlangıç tarihini kaydet
                    if operation_start_date is None:
                        operation_start_date = current_interval.get_date()

                    # Try to find consecutive intervals for the full operation
                    while assignment_duration < remaining_duration:
                        # Başlangıç tarihinden max_days_to_consider günden fazla ilerde mi kontrol et
                        days_difference = (current_interval.get_date() - operation_start_date).days
                        if days_difference > max_days_to_consider:
                            break

                        # Check if this operation is already assigned to this interval for this product
                        if self.is_operation_assigned_to_interval(operation.get_name(), current_interval, product):
                            print(
                                f"Operation {operation.get_name()} already assigned to interval {current_interval.get_date()} {current_interval.interval[0]}-{current_interval.interval[1]}")
                            # Move to next interval
                            next_interval = self.get_next_interval(current_interval, filtered_intervals)
                            if next_interval is None:
                                break
                            current_interval = next_interval
                            continue

                        # Verify this interval passes all constraints - pass product context
                        if (not self.previous_operation_control(operation, current_interval, product) or
                                (self.same_product_control(product, current_interval) and
                                 not self.check_jig_capacity(product, operation, current_interval)) or
                                not self.compatible_worker_number_check(operation, current_interval)):
                            break

                        # Select new workers if shift changes or for new interval
                        if current_interval.get_shift() != current_shift or current_interval not in interval_worker_assignments:
                            current_shift = current_interval.get_shift()
                            # Önce daha önce bu operasyonda çalışmış işçileri tercih et
                            selected_workers = self.select_best_workers_for_assignment(operation, current_interval,
                                                                                       prefer_previous=True)
                            if not selected_workers or len(selected_workers) < operation.get_required_worker():
                                break  # Not enough suitable workers for this interval
                            interval_worker_assignments[current_interval] = selected_workers

                        assignment_intervals.append(current_interval)
                        assignment_duration += 0.25  # Each interval is 0.25 hours

                        # Look for next interval
                        next_interval = self.get_next_interval(current_interval, filtered_intervals)
                        if next_interval is None:
                            break
                        current_interval = next_interval

                    # Check if we found enough intervals for full assignment
                    if assignment_intervals and assignment_duration > 0:
                        print(
                            f"Creating assignments for operation {operation.get_name()} of product {product.get_serial_number()}, duration: {assignment_duration} of {remaining_duration}")

                        # Create assignments for each interval with their corresponding workers
                        assignments_created = 0
                        for assigned_interval in assignment_intervals:
                            assigned_workers = interval_worker_assignments[assigned_interval]
                            if self.create_assignment(assigned_interval, jig, product, operation, assigned_workers):
                                assignments_created += 1

                        if assignments_created == 0:
                            print(
                                f"Failed to create any assignments for operation {operation.get_name()} of product {product.get_serial_number()}")
                            continue  # Try another interval

                        # Update remaining duration based on the actual assignments created
                        actual_assigned_duration = assignments_created * 0.25
                        new_remaining = remaining_duration - actual_assigned_duration
                        operation.set_remaining_duration(new_remaining)

                        # Mark operation as completed if full duration scheduled
                        if new_remaining <= 0.001:  # Use small epsilon for floating point comparison
                            operation.set_completed(True)
                            operation.set_remaining_duration(0)
                            successfully_assigned = True
                            print(
                                f"Operation {operation.get_name()} of product {product.get_serial_number()} fully assigned and completed")
                            break  # Break the interval loop
                        else:
                            # Partial assignment - keep looking for more intervals
                            print(
                                f"Partial assignment for operation {operation.get_name()} of product {product.get_serial_number()}, remaining: {new_remaining} hours")
                            # If we've assigned some time but not all, continue with next interval
                            if (assignment_duration < remaining_duration) and (assignments_created > 0):
                                remaining_duration = new_remaining
                                continue  # Try to assign more to this operation

                    # If we've assigned enough intervals for complete operation, stop looking
                    if successfully_assigned:
                        break

                # If we couldn't assign this operation at all, log it
                if not successfully_assigned and operation.get_remaining_duration() > 0.001:
                    print(
                        f"Could not fully assign operation {operation.get_name()} for product {product.get_serial_number()}")

        # Continue with preparation for the next round of assignments
        self.make_assignment_preparetions()

    def get_worker_assignments(self):
        """
        Her işçi için atama verilerini toplar.

        Returns:
            dict: İşçi adı anahtarlı, her işçinin atamalarını içeren sözlük
        """
        worker_assignments = {}

        # Tüm zaman aralıklarını kontrol et
        for date_obj in self.__ScheduleObject.dates:
            for time_interval in date_obj.time_intervals:
                for assignment in time_interval.get_assignments():
                    jig, product, operation, workers = assignment

                    # Zaman bilgilerini oluştur
                    assignment_date = date_obj.date.strftime("%d.%m.%Y")
                    start_time = time_interval.interval[0].strftime("%H:%M")
                    end_time = time_interval.interval[1].strftime("%H:%M")
                    time_range = f"{start_time}-{end_time}"

                    # Her işçi için atama bilgisini ekle
                    for worker in workers:
                        worker_name = worker.get_name()
                        reg_number = worker.get_registration_number()

                        # İşçi için girdi yoksa ekle
                        if worker_name not in worker_assignments:
                            worker_assignments[worker_name] = {
                                "registration_number": reg_number,
                                "assignments": []
                            }

                        # Atama bilgisini ekle
                        worker_assignments[worker_name]["assignments"].append({
                            "date": assignment_date,
                            "shift": time_interval.shift,
                            "time": time_range,
                            "product": product.get_serial_number(),
                            "operation": operation.get_name(),
                            "jig": jig.get_name()
                        })

        return worker_assignments
    def get_next_interval(self, current_interval, intervals_list):
        """
        Mevcut interval'dan sonraki interval'ı bulur.
        """
        current_index = intervals_list.index(current_interval)
        if current_index + 1 < len(intervals_list):
            return intervals_list[current_index + 1]
        return None  # Sonraki interval yok

    def select_best_workers_for_assignment(self, operation, time_interval, prefer_previous=True):
        """
        Operasyon için en uygun işçileri seçer.
        Daha önce bu operasyonda çalışmış işçileri tercih eder (eğer vardiyaları uygunsa).

        Args:
            operation: Atanacak operasyon
            time_interval: Zaman aralığı
            prefer_previous: Daha önce bu operasyonda çalışmış işçileri tercih et

        Returns:
            list: Seçilen işçilerin listesi
        """
        required_skills = operation.get_required_skills()
        required_worker_count = operation.get_required_worker()

        # Tüm uygun beceri ve vardiyaya sahip işçileri bul
        qualified_workers = []

        # Önce daha önce bu operasyona atanmış işçileri kontrol et
        previous_workers = []
        if prefer_previous:
            previous_workers = self.get_previous_workers_for_operation(operation.get_name())

        # Eğer önceki işçiler varsa ve vardiyaları uygunsa, onları tercih et
        previous_qualified_workers = []

        # Önce önceki işçileri kontrol et
        for worker in previous_workers:
            # Bu işçi mevcut zaman aralığında uygun mu kontrol et
            is_available = False
            for available_worker in time_interval.available_workers:
                if worker.get_registration_number() == available_worker.get_registration_number():
                    is_available = True
                    worker = available_worker  # Güncel işçi nesnesini kullan
                    break

            if not is_available:
                continue

            # Beceri kontrolü
            worker_skills = worker.get_skills()
            is_qualified = False

            if worker_skills in SKILLS:
                worker_skill_set = SKILLS[worker_skills]
                if required_skills in worker_skill_set:
                    is_qualified = True
            else:
                if required_skills == worker_skills:
                    is_qualified = True

            if is_qualified:
                # Vardiya kontrolü
                worker_shift_schedule = worker.get_shift_schedule()
                for schedule_entry in worker_shift_schedule:
                    schedule_date, schedule_shift, available_hours = schedule_entry

                    # Tarih ve vardiya kontrolü
                    if schedule_date == time_interval.get_date() and schedule_shift == time_interval.get_shift():
                        # İşçi hem beceri hem vardiya açısından uygun
                        previous_qualified_workers.append(worker)
                        break

        # Eğer tüm önceki işçiler mevcut aralık için uygunsa ve gerekli sayıda varsa onları kullan
        if len(previous_qualified_workers) >= required_worker_count:
            # İşçilerin atama sayılarını kontrol et, eğer atama sayısı tanımlı değilse 0 olarak başlat
            for worker in previous_qualified_workers:
                if not hasattr(worker, 'assignment_count'):
                    worker.assignment_count = 0

            # Sadece gerekli sayıda işçiyi seç
            return previous_qualified_workers[:required_worker_count]

        # Önceki işçiler yeterli değilse, tüm uygun işçileri bul
        for worker in time_interval.available_workers:
            # Eğer bu işçi zaten previous_qualified_workers içindeyse atla
            already_selected = False
            for prev_worker in previous_qualified_workers:
                if worker.get_registration_number() == prev_worker.get_registration_number():
                    already_selected = True
                    break

            if already_selected:
                continue

            # Beceri kontrolü
            worker_skills = worker.get_skills()
            is_qualified = False

            if worker_skills in SKILLS:
                worker_skill_set = SKILLS[worker_skills]
                if required_skills in worker_skill_set:
                    is_qualified = True
            else:
                if required_skills == worker_skills:
                    is_qualified = True

            if is_qualified:
                # Vardiya kontrolü
                worker_shift_schedule = worker.get_shift_schedule()
                for schedule_entry in worker_shift_schedule:
                    schedule_date, schedule_shift, available_hours = schedule_entry

                    # Tarih ve vardiya kontrolü
                    if schedule_date == time_interval.get_date() and schedule_shift == time_interval.get_shift():
                        # İşçi hem beceri hem vardiya açısından uygun
                        qualified_workers.append(worker)
                        break

        # Tüm uygun işçileri birleştir (önce önceki işçiler)
        all_qualified_workers = previous_qualified_workers + qualified_workers

        # Yeterli sayıda uygun işçi var mı kontrol et
        if len(all_qualified_workers) < required_worker_count:
            return None

        # İşçilerin atama sayılarını kontrol et, eğer atama sayısı tanımlı değilse 0 olarak başlat
        for worker in all_qualified_workers:
            if not hasattr(worker, 'assignment_count'):
                worker.assignment_count = 0

        # İşçileri atama sayısına göre sırala (en az atama yapılan önce)
        # Ama önceki işçileri önde tut
        sorted_workers = sorted(all_qualified_workers,
                                key=lambda w: (0 if w in previous_qualified_workers else 1, w.assignment_count))

        # Gerekli sayıda işçiyi seç
        selected_workers = sorted_workers[:required_worker_count]

        return selected_workers

    def find_latest_finish_time_of_predecessors(self, operation, product):
        """
        Finds the latest finish time of all predecessors for an operation within a specific product.

        Args:
            operation: The operation to check predecessors for
            product: The product this operation belongs to (required parameter)

        Returns:
            A tuple of (date, time) representing the latest finish time, or None if any predecessor hasn't been scheduled
        """
        op = operation
        latest_finish_time = None
        intervals = self.get_ScheduleObject().get_sorted_time_intervals()

        # If no predecessors, can start at the earliest possible time
        if len(op.get_predecessors()) == 0:
            first_interval = intervals[0]
            latest_finish_time = (first_interval.get_date(), first_interval.interval[0])
            return latest_finish_time

        # Only consider predecessors from the same product
        all_predecessors_have_end_time = True
        product_predecessors = []

        # Get all valid predecessors for this product
        for prev_op in op.get_predecessors():
            # Ensure the predecessor belongs to this product
            if prev_op in product.get_operations():
                product_predecessors.append(prev_op)
                # Check if the predecessor has an end time (has been scheduled)
                if not prev_op.get_end_datetime() and not prev_op.get_completed():
                    print(
                        f"Predecessor {prev_op.get_name()} for operation {op.get_name()} of product {product.get_serial_number()} has no end time and is not completed")
                    all_predecessors_have_end_time = False
                    break

        # If any valid predecessor doesn't have an end time, this operation can't be scheduled yet
        if not all_predecessors_have_end_time:
            return None

        # Find the latest end time among all valid predecessors
        for prev_op in product_predecessors:
            # Skip if the predecessor is already marked as completed but doesn't have end_datetime
            if prev_op.get_completed() and not prev_op.get_end_datetime():
                continue

            # If predecessor has an end_datetime, compare it with current latest
            if prev_op.get_end_datetime():
                if latest_finish_time is None:
                    latest_finish_time = prev_op.get_end_datetime()
                elif prev_op.get_end_datetime()[0] > latest_finish_time[0]:
                    latest_finish_time = prev_op.get_end_datetime()
                elif (prev_op.get_end_datetime()[0] == latest_finish_time[0] and
                      prev_op.get_end_datetime()[1] > latest_finish_time[1]):
                    latest_finish_time = prev_op.get_end_datetime()

        return latest_finish_time

    # 2. Fix the filter_intervals_after_time method to ensure operations start strictly after predecessors end
    def filter_intervals_after_time(self, intervals_list, start_time):
        if start_time is None:
            return []  # If no start time is found, don't schedule (empty list)

        # Tercih edilen vardiya tipini al (I1, I2, I3)
        preferred_shift = self.screenController.get_starting_shift()

        # İlk gün ve vardiyayı bul
        first_day = None
        if intervals_list:
            first_day = intervals_list[0].get_date()

        filtered_intervals = []

        # Eğer önceki operasyonlar daha önce tamamlanmışsa ve ilk gündeysek
        if start_time is None or (first_day and (start_time[0] < first_day or
                                                 (start_time[0] == first_day and start_time[1] <=
                                                  intervals_list[0].interval[0]))):
            # Önce ilk gün tercih edilen vardiyaya ait aralıkları ekle
            first_day_intervals = {shift: [] for shift in ["I1", "I2", "I3"]}
            other_intervals = []

            for interval in intervals_list:
                if interval.get_date() == first_day:
                    first_day_intervals[interval.get_shift()].append(interval)
                else:
                    other_intervals.append(interval)

            # Tercih edilen vardiyadan başlayarak, ilk gün için vardiyaları sırala
            shift_order = []
            if preferred_shift == "I1":
                shift_order = ["I1", "I2", "I3"]
            elif preferred_shift == "I2":
                shift_order = ["I2", "I3", "I1"]
            elif preferred_shift == "I3":
                shift_order = ["I3", "I1", "I2"]

            # İlk gün vardiyalarını tercih sırasına göre ekle
            for shift in shift_order:
                filtered_intervals.extend(first_day_intervals[shift])

            # Diğer günlerin aralıklarını ekle
            filtered_intervals.extend(other_intervals)

            return filtered_intervals

        # Normal durum: önceki operasyonların tamamlanma zamanından sonraki aralıkları filtrele
        for interval in intervals_list:
            interval_date = interval.get_date()
            interval_start_time = interval.interval[0]  # Interval's start time

            # Only include intervals that start strictly after the predecessor end time
            if interval_date > start_time[0]:
                filtered_intervals.append(interval)
            elif interval_date == start_time[0] and interval_start_time > start_time[1]:
                filtered_intervals.append(interval)

        return filtered_intervals

    def get_skill_priority(self, skill):
        priority_order = {
            "KISMİ": 1,
            "KALİTE": 2,
            "ÜRETİM DIŞI": 3,
            "ÜRETİM": 4,
            "HEPSİ": 5
        }
        return priority_order.get(skill, 6)  # Eğer skill tanımlı değilse en düşük öncelik



    def same_product_control(self, product, time_interval):  # Buraya operasyonun ait olduğu product gönderilecek
        assignments = time_interval.get_assignments()
        for assignment in assignments:
            if product == assignment[1]:
                return True  # aralıkta aynı product'a ait operasyon var
        return False  # aralıkta aynı product'a ait operasyonn yok

    def jig_compatibility_control(self, product, operation):
        if product.get_current_jig() in operation.get_compatible_jigs():
            return True
        else:
            return False

    def change_jig(self, product, operation):
        jig_names = operation.get_compatible_jigs()
        jig_object_list = []
        for jig_name in jig_names:
            jig_object_list.append(self.get_jig(jig_name))
        for jig in jig_object_list:
            if not jig.get_state():
                jig.set_state(True)
                product.set_current_jig(jig)
            else:
                break

    def check_jig_capacity(self, product, operation, time_interval):
        """
        Checks if there's enough jig capacity for the operation in the given time interval.
        Also prevents assigning the same operation multiple times.
        """
        jig = product.get_current_jig()
        total_workers_assigned = 0
        operation_already_assigned = False

        # Bu aralıktaki mevcut atamaları kontrol et
        for assignment in time_interval.get_assignments():
            assigned_jig, assigned_product, assigned_operation, assigned_workers = assignment

            # Bu operasyon zaten bu aralıkta atanmışsa, tekrar atama
            if assigned_operation.get_name() == operation.get_name():
                print(f"Operation {operation.get_name()} already assigned in check_jig_capacity")
                return False

            # Atama aynı jig ve aynı ürüne aitse, işçi sayısını ekle
            if assigned_jig == jig and assigned_product == product:
                total_workers_assigned += len(assigned_workers)

        # Yeni operasyon için gereken işçi sayısını ekleyerek toplamı kontrol et
        if total_workers_assigned + operation.get_required_worker() <= jig.get_max_assigned_worker():
            return True  # Jig kapasitesi aşılmıyor, atama yapılabilir
        else:
            print(
                f"Jig capacity exceeded: {total_workers_assigned} + {operation.get_required_worker()} > {jig.get_max_assigned_worker()}")
            return False  # Jig kapasitesi aşılıyor, atama yapılamaz

    def compatible_worker_number_check(self, operation, time_interval):
        required_skills = operation.get_required_skills()
        available_workers = []

        for w in time_interval.available_workers:
            worker_skills = w.get_skills()

            # worker_skills'i SKILLS sözlüğünde key olarak ara
            if worker_skills in SKILLS:
                # worker_skills'in karşılığındaki yetenek kümesini al
                worker_skill_set = SKILLS[worker_skills]

                # required_skills, worker_skill_set içinde var mı kontrol et
                if required_skills in worker_skill_set:
                    available_workers.append(w)
            else:
                # worker_skills SKILLS sözlüğünde yoksa, doğrudan eşleşme kontrolü yap
                if required_skills == worker_skills:
                    available_workers.append(w)

        # Yeterli sayıda uyumlu çalışan var mı kontrol et
        if len(available_workers) >= operation.get_required_worker():
            return True  # Atama yapılabilir
        else:
            return False  # Atama yapılamaz

    def create_assignment(self, time_interval, jig, product, operation, workers):
        """
        Creates an assignment for an operation in a time interval.
        Updates operation completion status and timestamps.

        Args:
            time_interval: The time interval for the assignment
            jig: The jig to use
            product: The product being processed
            operation: The operation to perform
            workers: The workers assigned

        Returns:
            True if the assignment was created successfully, False otherwise
        """
        # Double-check that the operation can be scheduled
        if not self.previous_operation_control(operation, time_interval, product):
            print(
                f"WARNING: Operation {operation.get_name()} for product {product.get_serial_number()} cannot be scheduled in this interval - predecessor check failed")
            return False

        # Check if this operation is already assigned to this interval for this product
        if self.is_operation_assigned_to_interval(operation.get_name(), time_interval, product):
            print(
                f"WARNING: Operation {operation.get_name()} already assigned to this interval - skipping duplicate assignment")
            return False

        inter = time_interval
        assignment_entry = jig, product, operation, workers
        inter.set_assignments(assignment_entry)
        jig.set_state(True)
        product.set_current_jig(jig)

        # Operasyon için işçileri güncelle/ekle
        self.update_operation_workers(operation.get_name(), workers)

        # Operasyon atamasını takip et (eğer böyle bir sistem eklediyseniz)
        if hasattr(self, 'add_operation_assignment'):
            self.add_operation_assignment(operation.get_name(), time_interval, len(workers))

        # Kalan süreyi güncelle
        remaining = operation.get_remaining_duration() or operation.get_operating_duration()
        remaining -= 0.25  # Her zaman aralığı 0.25 gün
        operation.set_remaining_duration(remaining)

        # Sadece kalan süre 0 veya daha az ise tamamlandı olarak işaretle
        if remaining <= 0.001:  # Küçük bir epsilon değeri kullan
            operation.set_completed(True)
            print(f"Operation {operation.get_name()} completed with remaining duration: {remaining}")
        else:
            operation.set_completed(False)
            print(f"Operation {operation.get_name()} partially assigned, remaining: {remaining}")

        operation.set_start_datetime(inter.get_date(), inter.interval[0])
        operation.set_end_datetime(inter.get_date(), inter.interval[1])

        # İşçilerin atama sayısını artır
        for worker in workers:
            if not hasattr(worker, 'assignment_count'):
                worker.assignment_count = 0
            worker.assignment_count += 1

        self.update_worker_shift_schedule(workers, inter)
        self.assign_workers_to_time_intervals()

        return True  # Başarılı atama


    def update_worker_shift_schedule(self, workers, time_interval):
        """
        Atama yapılan çalışanların shift schedule'ını günceller.
        Atama yapılan zaman aralığını çalışanların schedule'ından çıkarır.
        """
        interval_start = time_interval.interval[0]  # Zaman aralığının başlangıcı
        interval_end = time_interval.interval[1]  # Zaman aralığının bitişi
        interval_date = time_interval.get_date()  # Zaman aralığının tarihi
        interval_shift = time_interval.shift  # Zaman aralığının vardiyası

        for worker in workers:
            # Çalışanın shift schedule'ını al
            shift_schedule = worker.get_shift_schedule()

            # Shift schedule'ı güncelle
            updated_schedule = []
            for schedule_entry in shift_schedule:
                schedule_date, schedule_shift, available_hours = schedule_entry

                # Eğer tarih ve vardiya eşleşiyorsa, zaman aralığını çıkar
                if schedule_date == interval_date and schedule_shift == interval_shift:
                    new_available_hours = []
                    for hours in available_hours:
                        # Zaman aralığını çıkar
                        if not (hours[0] <= interval_start and hours[1] >= interval_end):
                            new_available_hours.append(hours)

                    # Eğer yeni available_hours boş değilse, schedule'a ekle
                    if new_available_hours:
                        updated_schedule.append((schedule_date, schedule_shift, new_available_hours))
                else:
                    # Tarih ve vardiya eşleşmiyorsa, schedule'ı olduğu gibi ekle
                    updated_schedule.append(schedule_entry)

            # Çalışanın shift schedule'ını güncelle
            worker.set_shift_schedule(updated_schedule)

    def update_operation_remaining_duration(self, operation, completion_percentage):
        """
        Bir operasyonun kalan süresini ilerleme yüzdesine göre günceller

        Args:
            operation: Güncellenecek operasyon
            completion_percentage: Tamamlanma yüzdesi (0-100 arası)
        """
        if completion_percentage < 0:
            completion_percentage = 0
        elif completion_percentage > 100:
            completion_percentage = 100

        # Toplam operasyon süresini al
        total_duration = operation.get_operating_duration()

        # Kalan süreyi hesapla
        remaining_duration = total_duration * (1 - completion_percentage / 100)

        # Kalan süreyi güncelle
        operation.set_remaining_duration(remaining_duration)

        # Eğer tam tamamlandıysa, completed olarak işaretle
        if completion_percentage == 100:
            operation.set_completed(True)
        else:
            operation.set_completed(False)

        print(
            f"Operation {operation.get_name()} progress: {completion_percentage}%, remaining duration: {remaining_duration}")

    def debug_operation_durations(self, serial_number):
        """Debug function to trace operation durations"""
        product = self.get_product(serial_number)
        print(f"\n===== DEBUG: Operation durations for product {serial_number} =====")
        for op in product.get_operations():
            print(f"Operation {op.get_name()}:")
            print(f"  * Full duration: {op.get_operating_duration()}")
            print(f"  * Remaining duration: {op.get_remaining_duration()}")
            print(f"  * Completed: {op.get_completed()}")
            print(f"  * Required worker: {op.get_required_worker()}")
            print(f"  * Required man hours: {op.get_required_man_hours()}")
        print("===========================================\n")

    def make_assignment_preparetions(self):
        """
        Prepares for the next round of assignments by updating product progress,
        recalculating critical operations, and initiating new assignments.
        Also handles consistency checks and deduplication of operations.
        """
        # Yeni bir atama başlamadan önce atama takibini temizle (eğer böyle bir sistem eklediyseniz)
        if hasattr(self, 'clear_assignments_tracking'):
            self.clear_assignments_tracking()

        plist = self.__products

        # Her ürün için tutarlılık kontrolü
        for product in plist:
            for op in product.get_operations():
                # Kalan süreyi kontrol et
                remaining = op.get_remaining_duration()
                completed = op.get_completed()

                # Tutarsızlık kontrolü
                if remaining is not None and remaining > 0.001 and completed:
                    print(
                        f"Fixing inconsistency in op {op.get_name()}: marked completed but has remaining time {remaining}")
                    op.set_completed(False)  # Completed flag'i düzelt

                # Eğer kalan süre tanımlanmamışsa, tam süreyi kullan
                if remaining is None and not completed:
                    op.set_remaining_duration(op.get_operating_duration())

        # Debug için her ürünün operasyon sürelerini yazdır
        for product in plist:
            self.debug_operation_durations(product.get_serial_number())

        # Ürünlerin ilerlemesini hesapla ve öncülleri güncelle
        for product in plist:
            sn = product.get_serial_number()
            self.calculate_product_progress(sn)
            self.remove_completed_predecessors(sn)
            self.set_critical_operations(sn)

        # Operasyonları süreye göre sırala
        self.sort_operations_by_duration()
        # Ürünleri ilerleme durumuna göre sırala
        self.sort_products_by_progress()

        # Yeni kritik operasyon listesi
        new_critical_ops = self.get_all_critical_operations()

        # Aynı ürün-operasyon çiftinin tekrarlanmasını önle
        unique_critical_ops = []
        seen_product_ops = set()  # (product_serial, op_name) çiftlerini takip etmek için

        for product, op in new_critical_ops:
            product_op_key = (product.get_serial_number(), op.get_name())
            if product_op_key not in seen_product_ops:
                seen_product_ops.add(product_op_key)
                unique_critical_ops.append((product, op))
            else:
                print(
                    f"Skipping duplicate critical operation: {op.get_name()} for product {product.get_serial_number()}")

        # Sonsuz döngü kontrolü
        if unique_critical_ops == self.__critical_op_check_list and unique_critical_ops:
            print("Same operation list detected in consecutive iterations - may be stuck in a loop")
            # Eğer aynı operasyon listesi tekrar ediyorsa, ilk operasyonu zorla çözmeye çalış
            if unique_critical_ops:
                product, operation = unique_critical_ops[0]
                print(
                    f"Forcing resolution for operation {operation.get_name()} of product {product.get_serial_number()}")
                # Operasyonu tamamlandı olarak işaretle ve atama listesinden çıkar
                operation.set_completed(True)
                operation.set_remaining_duration(0)
                # Yeni bir liste oluştur ve devam et
                self.__critical_op_check_list = []
                self.make_assignment_preparetions()
                return

        # Aktif listeyi güncelle (döngü tespiti için)
        self.__critical_op_check_list = unique_critical_ops

        # Eğer kritik operasyon listesi boşsa, işlem tamamlandı demektir
        if not unique_critical_ops:
            print("Assignment complete - no more critical operations.")
            return

        # Debug için kritik operasyonları yazdır
        print("Critical operations for next assignment:")
        for product, op in unique_critical_ops:
            print(
                f"Product: {product.get_serial_number()}, Operation: {op.get_name()}, Remaining: {op.get_remaining_duration()}")

        # Atama işlemini başlat - tekrarlanan operasyonları çıkardıktan sonra
        self.initiate_assignment(unique_critical_ops)
    def get_assignments_for_output(self):
        assignments = []
        for date_obj in self.__ScheduleObject.dates:
            for time_interval in date_obj.time_intervals:
                for assignment in time_interval.get_assignments():
                    jig, product, operation, workers = assignment
                    # Time interval'ları birleştirerek kapsayıcı bir aralık oluştur
                    start_time = time_interval.interval[0].strftime("%H:%M")
                    end_time = time_interval.interval[1].strftime("%H:%M")
                    time_range = f"{start_time}-{end_time}"

                    # Workers'ları birleştirerek bir string oluştur
                    worker_names = ", ".join([worker.get_name() for worker in workers])

                    # Atamayı listeye ekle
                    assignments.append({
                        "Product": product.get_serial_number(),
                        "Jig": jig.get_name(),
                        "Operation": operation.get_name(),
                        "Date": date_obj.date.strftime("%d.%m.%Y"),
                        "Shift": time_interval.shift,
                        "Time Interval": time_range,
                        "Workers": worker_names
                    })
        return assignments

    def export_assignments_to_excel(self, file_path=None):
        """
        Atama çıktılarını Excel dosyasına aktarır.
        Her ürün için ayrı bir sayfa, her işçi için ayrı bir sayfa ve bir Gantt şeması oluşturur.

        :param file_path: Excel dosyasının kaydedileceği konum. None ise kullanıcıdan sorulur.
        :return: Başarılı olursa True, aksi halde False
        """
        try:
            # Atama çıktılarını al
            assignments = self.get_assignments_for_output()
            worker_assignments = self.get_worker_assignments()

            if not assignments:
                print("No assignments to export.")
                return False

            # Eğer dosya yolu belirtilmemişse, kullanıcıdan al
            if not file_path:
                from tkinter import filedialog
                file_path = filedialog.asksaveasfilename(
                    title="Save Excel File",
                    filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")),
                    defaultextension=".xlsx"
                )

                if not file_path:  # Kullanıcı iptal ettiyse
                    return False

            # Yeni bir Excel çalışma kitabı oluştur
            wb = openpyxl.Workbook()
            # Varsayılan sayfayı sil
            wb.remove(wb.active)

            # Stil tanımlamaları
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # --- ÜRÜN BAZLI SAYFALAR ---
            product_assignments = {}
            for assignment in assignments:
                product_serial = assignment["Product"]
                if product_serial not in product_assignments:
                    product_assignments[product_serial] = []
                product_assignments[product_serial].append(assignment)

            # Her ürün için ayrı bir sayfa oluştur
            for product_serial, product_assignments_list in product_assignments.items():
                # Operasyon adına göre sırala
                sorted_assignments = sorted(product_assignments_list,
                                            key=lambda x: int(x["Operation"]) if x["Operation"].isdigit() else x[
                                                "Operation"])

                # Yeni sayfa oluştur
                sheet = wb.create_sheet(f"Product {product_serial}")

                # Başlıkları ayarla
                headers = ["Operation", "Jig", "Date", "Shift", "Time Interval", "Workers"]
                for col_idx, header in enumerate(headers, 1):
                    cell = sheet.cell(row=1, column=col_idx)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

                    # Sütun genişliklerini ayarla
                    if header == "Workers":
                        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 30
                    elif header == "Time Interval":
                        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
                    else:
                        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12

                # Verileri doldur
                for row_idx, assignment in enumerate(sorted_assignments, 2):
                    sheet.cell(row=row_idx, column=1).value = assignment["Operation"]
                    sheet.cell(row=row_idx, column=2).value = assignment["Jig"]
                    sheet.cell(row=row_idx, column=3).value = assignment["Date"]
                    sheet.cell(row=row_idx, column=4).value = assignment["Shift"]
                    sheet.cell(row=row_idx, column=5).value = assignment["Time Interval"]
                    sheet.cell(row=row_idx, column=6).value = assignment["Workers"]

                    # Hücre stillerini ayarla
                    for col_idx in range(1, 7):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        cell.alignment = Alignment(horizontal='center')
                        if row_idx % 2 == 0:
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

                # Özet bilgiler
                row_idx = len(sorted_assignments) + 3
                sheet.cell(row=row_idx, column=1).value = "Total Operations:"
                sheet.cell(row=row_idx, column=2).value = len(sorted_assignments)
                sheet.cell(row=row_idx, column=1).font = Font(bold=True)

                # Sayfayı güzelleştir
                for row in sheet.iter_rows(min_row=1, max_row=row_idx, min_col=1, max_col=6):
                    for cell in row:
                        cell.border = thin_border

            # --- İŞÇİ BAZLI SAYFALAR ---
            # Her işçi için ayrı bir sayfa oluştur
            for worker_name, data in worker_assignments.items():
                # Güvenli bir sayfa adı oluştur (Excel 31 karakterle sınırlıdır)
                safe_name = worker_name[:25] if len(worker_name) > 25 else worker_name
                sheet_name = f"Worker {safe_name}"

                # Yeni sayfa oluştur
                sheet = wb.create_sheet(sheet_name)

                # İşçi bilgileri başlık satırı
                sheet.cell(row=1, column=1).value = "Worker Name:"
                sheet.cell(row=1, column=2).value = worker_name
                sheet.cell(row=2, column=1).value = "Registration No:"
                sheet.cell(row=2, column=2).value = data["registration_number"]

                # Stil ayarları
                sheet.cell(row=1, column=1).font = Font(bold=True)
                sheet.cell(row=2, column=1).font = Font(bold=True)

                # Atama tablosu başlıkları
                headers = ["Date", "Shift", "Time", "Product", "Operation", "Jig"]
                for col_idx, header in enumerate(headers, 1):
                    cell = sheet.cell(row=4, column=col_idx)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                    sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12

                # Atamaları tarihe ve zamana göre sırala
                sorted_assignments = sorted(
                    data["assignments"],
                    key=lambda x: (
                        datetime.strptime(x["date"], "%d.%m.%Y"),
                        x["time"].split("-")[0]
                    )
                )

                # Verileri doldur
                for row_idx, assignment in enumerate(sorted_assignments, 5):
                    sheet.cell(row=row_idx, column=1).value = assignment["date"]
                    sheet.cell(row=row_idx, column=2).value = assignment["shift"]
                    sheet.cell(row=row_idx, column=3).value = assignment["time"]
                    sheet.cell(row=row_idx, column=4).value = assignment["product"]
                    sheet.cell(row=row_idx, column=5).value = assignment["operation"]
                    sheet.cell(row=row_idx, column=6).value = assignment["jig"]

                    # Hücre stillerini ayarla
                    for col_idx in range(1, 7):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        cell.alignment = Alignment(horizontal='center')
                        if row_idx % 2 == 0:
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        cell.border = thin_border

                # Özet bilgiler
                row_idx = len(sorted_assignments) + 6
                sheet.cell(row=row_idx, column=1).value = "Total Assignments:"
                sheet.cell(row=row_idx, column=2).value = len(sorted_assignments)
                sheet.cell(row=row_idx, column=1).font = Font(bold=True)

                # Sayfayı güzelleştir
                for row in sheet.iter_rows(min_row=1, max_row=row_idx, min_col=1, max_col=6):
                    for cell in row:
                        if not cell.border:
                            cell.border = thin_border

            # --- GANTT CHART SAYFASI ---
            gantt_sheet = wb.create_sheet("Worker Gantt Chart")

            # Tüm tarihleri topla ve sırala
            all_dates = set()
            for assignment in assignments:
                all_dates.add(assignment["Date"])
            sorted_dates = sorted(list(all_dates), key=lambda x: datetime.strptime(x, "%d.%m.%Y"))

            # Tüm zaman aralıklarını topla
            all_time_intervals = set()
            for assignment in assignments:
                all_time_intervals.add(assignment["Time Interval"])

            # Zaman aralıklarını saatlere göre sırala
            sorted_time_intervals = sorted(
                list(all_time_intervals),
                key=lambda x: datetime.strptime(x.split("-")[0], "%H:%M").time()
            )

            # Başlık satırı
            gantt_sheet.cell(row=1, column=1).value = "Worker Gantt Chart"
            gantt_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
            gantt_sheet.merge_cells('A1:E1')
            gantt_sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')

            # Alt başlık - İşçi bilgileri başlıkları
            gantt_sheet.cell(row=3, column=1).value = "Worker Name"
            gantt_sheet.cell(row=3, column=2).value = "Registration No"
            gantt_sheet.cell(row=3, column=3).value = "Total Hours"

            # Sütun genişliklerini ayarla
            gantt_sheet.column_dimensions['A'].width = 25  # İşçi adı sütunu genişliği
            gantt_sheet.column_dimensions['B'].width = 15  # Sicil no sütunu genişliği
            gantt_sheet.column_dimensions['C'].width = 12  # Toplam saat sütunu genişliği

            # Stil ayarlamaları - başlıklar
            for col in range(1, 4):
                cell = gantt_sheet.cell(row=3, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                cell.border = thin_border

            # Tarih ve saat başlıkları
            col_offset = 4  # İlk 3 sütun işçi bilgileri için ayrıldı
            for date_idx, date_str in enumerate(sorted_dates):
                date_col = col_offset + date_idx * (len(sorted_time_intervals) + 1)
                date_obj = datetime.strptime(date_str, "%d.%m.%Y")
                date_display = date_obj.strftime("%d.%m.%Y")

                # Tarih başlığı
                gantt_sheet.cell(row=2, column=date_col).value = date_display
                gantt_sheet.cell(row=2, column=date_col).font = Font(bold=True)
                gantt_sheet.merge_cells(
                    start_row=2,
                    start_column=date_col,
                    end_row=2,
                    end_column=date_col + len(sorted_time_intervals) - 1
                )
                gantt_sheet.cell(row=2, column=date_col).alignment = Alignment(horizontal='center')

                # Saat aralığı başlıkları
                for time_idx, time_interval in enumerate(sorted_time_intervals):
                    time_col = date_col + time_idx
                    gantt_sheet.cell(row=3, column=time_col).value = time_interval
                    gantt_sheet.cell(row=3, column=time_col).font = Font(bold=True)
                    gantt_sheet.cell(row=3, column=time_col).alignment = Alignment(horizontal='center')
                    gantt_sheet.column_dimensions[openpyxl.utils.get_column_letter(time_col)].width = 10

            # Stil ayarlamaları - tarih ve saat başlıkları
            for row in range(2, 4):
                for col in range(4, col_offset + len(sorted_dates) * (len(sorted_time_intervals) + 1)):
                    cell = gantt_sheet.cell(row=row, column=col)
                    if cell.value:  # Boş olmayan hücreler için
                        cell.border = thin_border
                        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

            # İşçileri ve atamalarını ekle
            row_offset = 4
            worker_colors = {}  # İşçi renklerini takip etmek için
            color_index = 0
            color_palette = [
                "9BC2E6",  # Açık mavi
                "A9D08E",  # Açık yeşil
                "FFD966",  # Sarı
                "F4B084",  # Turuncu
                "C9C9C9",  # Gri
                "8EA9DB",  # Mavi
                "FF99CC",  # Pembe
                "AEAAAA",  # Gümüş
            ]

            # Her işçi için toplam çalışma saatlerini hesapla
            worker_total_hours = {}
            for worker_name, data in worker_assignments.items():
                total_hours = 0

                for assignment in data["assignments"]:
                    # Zaman aralığını parçala
                    time_str = assignment["time"]
                    start_time_str, end_time_str = time_str.split("-")

                    # Saat ve dakikayı al
                    start_time = datetime.strptime(start_time_str, "%H:%M").time()
                    end_time = datetime.strptime(end_time_str, "%H:%M").time()

                    # Saat farkını hesapla (saatler ve dakikalar)
                    start_minutes = start_time.hour * 60 + start_time.minute
                    end_minutes = end_time.hour * 60 + end_time.minute

                    # Gece yarısı geçişini kontrol et
                    if end_minutes < start_minutes:  # Örneğin 23:00-01:00 durumu
                        end_minutes += 24 * 60  # Bir günlük dakikaları ekle

                    # Dakika farkını saate çevir
                    hours_diff = (end_minutes - start_minutes) / 60
                    total_hours += hours_diff

                worker_total_hours[worker_name] = total_hours

            # Tüm işçiler için sırayla satır oluştur
            for worker_name, data in worker_assignments.items():
                worker_row = row_offset
                row_offset += 1

                # İşçi adı
                gantt_sheet.cell(row=worker_row, column=1).value = worker_name
                gantt_sheet.cell(row=worker_row, column=1).alignment = Alignment(horizontal='left')
                gantt_sheet.cell(row=worker_row, column=1).border = thin_border

                # Sicil numarası
                gantt_sheet.cell(row=worker_row, column=2).value = data["registration_number"]
                gantt_sheet.cell(row=worker_row, column=2).alignment = Alignment(horizontal='center')
                gantt_sheet.cell(row=worker_row, column=2).border = thin_border

                # Toplam çalışma saati - doğru hesaplanmış hali
                total_hours = worker_total_hours[worker_name]
                gantt_sheet.cell(row=worker_row, column=3).value = round(total_hours, 2)  # 2 ondalık basamağa yuvarla
                gantt_sheet.cell(row=worker_row, column=3).alignment = Alignment(horizontal='center')
                gantt_sheet.cell(row=worker_row, column=3).border = thin_border

                # Bu işçi için renk seç
                if worker_name not in worker_colors:
                    worker_colors[worker_name] = color_palette[color_index % len(color_palette)]
                    color_index += 1

                worker_color = worker_colors[worker_name]

                # İşçinin atamalarını işle
                for assignment in data["assignments"]:
                    date_str = assignment["date"]
                    time_interval = assignment["time"]
                    operation = assignment["operation"]
                    product = assignment["product"]

                    # Tarihin indeksini bul
                    date_idx = sorted_dates.index(date_str)
                    # Zaman aralığının indeksini bul
                    time_idx = sorted_time_intervals.index(time_interval)

                    # Gantt hücresinin konumunu hesapla
                    cell_col = col_offset + date_idx * (len(sorted_time_intervals) + 1) + time_idx

                    # Hücreyi doldur
                    cell = gantt_sheet.cell(row=worker_row, column=cell_col)
                    cell.value = f"{product}-{operation}"
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color=worker_color, end_color=worker_color, fill_type="solid")
                    cell.border = thin_border

            # Açıklayıcı not ekle
            note_row = row_offset + 2
            gantt_sheet.cell(row=note_row, column=1).value = "Note: Each cell contains 'Product-Operation'"
            gantt_sheet.cell(row=note_row, column=1).font = Font(italic=True)
            gantt_sheet.merge_cells(
                start_row=note_row,
                start_column=1,
                end_row=note_row,
                end_column=5
            )

            # Boş hücrelere ince kenarlık ekle
            for row in range(4, row_offset):
                for col in range(4, col_offset + len(sorted_dates) * (len(sorted_time_intervals) + 1)):
                    cell = gantt_sheet.cell(row=row, column=col)
                    if not cell.value:  # Boş hücreler için
                        cell.border = thin_border

            # --- ÖZET SAYFASI ---
            summary_sheet = wb.create_sheet("Summary", 0)  # İlk sayfaya yerleştir
            summary_sheet.cell(row=1, column=1).value = "Assignment Summary"
            summary_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
            summary_sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')
            summary_sheet.merge_cells('A1:D1')

            summary_sheet.cell(row=3, column=1).value = "Product"
            summary_sheet.cell(row=3, column=2).value = "Operations Count"
            summary_sheet.cell(row=3, column=3).value = "First Date"
            summary_sheet.cell(row=3, column=4).value = "Last Date"

            for col_idx in range(1, 5):
                summary_sheet.cell(row=3, column=col_idx).font = Font(bold=True)
                summary_sheet.cell(row=3, column=col_idx).alignment = Alignment(horizontal='center')
                summary_sheet.cell(row=3, column=col_idx).fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7",
                                                                             fill_type="solid")
                summary_sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15

            row_idx = 4
            for product_serial, product_assignments_list in product_assignments.items():
                first_date = min(product_assignments_list, key=lambda x: datetime.strptime(x["Date"], "%d.%m.%Y"))[
                    "Date"]
                last_date = max(product_assignments_list, key=lambda x: datetime.strptime(x["Date"], "%d.%m.%Y"))[
                    "Date"]

                summary_sheet.cell(row=row_idx, column=1).value = product_serial
                summary_sheet.cell(row=row_idx, column=2).value = len(product_assignments_list)
                summary_sheet.cell(row=row_idx, column=3).value = first_date
                summary_sheet.cell(row=row_idx, column=4).value = last_date

                # Hücre stillerini ayarla
                for col_idx in range(1, 5):
                    cell = summary_sheet.cell(row=row_idx, column=col_idx)
                    cell.alignment = Alignment(horizontal='center')
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    cell.border = thin_border

                row_idx += 1

            # İşçi özeti
            row_idx += 2  # Boşluk bırak
            summary_sheet.cell(row=row_idx, column=1).value = "Worker Summary"
            summary_sheet.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
            summary_sheet.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center')
            summary_sheet.merge_cells(f'A{row_idx}:D{row_idx}')

            row_idx += 1
            summary_sheet.cell(row=row_idx, column=1).value = "Worker Name"
            summary_sheet.cell(row=row_idx, column=2).value = "Registration Number"
            summary_sheet.cell(row=row_idx, column=3).value = "Assignments Count"
            summary_sheet.cell(row=row_idx, column=4).value = "Total Hours"

            for col_idx in range(1, 5):
                summary_sheet.cell(row=row_idx, column=col_idx).font = Font(bold=True)
                summary_sheet.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center')
                summary_sheet.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color="DDEBF7",
                                                                                   end_color="DDEBF7",
                                                                                   fill_type="solid")

            row_idx += 1
            for worker_name, data in worker_assignments.items():
                assignments_count = len(data["assignments"])
                # Düzgün hesaplanmış toplam saat
                total_hours = worker_total_hours[worker_name]

                summary_sheet.cell(row=row_idx, column=1).value = worker_name
                summary_sheet.cell(row=row_idx, column=2).value = data["registration_number"]
                summary_sheet.cell(row=row_idx, column=3).value = assignments_count
                summary_sheet.cell(row=row_idx, column=4).value = round(total_hours, 2)  # 2 ondalık basamağa yuvarla

                # Hücre stillerini ayarla
                for col_idx in range(1, 5):
                    cell = summary_sheet.cell(row=row_idx, column=col_idx)
                    cell.alignment = Alignment(horizontal='center')
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    cell.border = thin_border

                row_idx += 1

            # Dosyayı kaydet
            wb.save(file_path)
            print(f"Assignments exported to {file_path}")
            return True

        except Exception as e:
            print(f"Error exporting assignments to Excel: {e}")
            import traceback
            traceback.print_exc()
            return False

    def export_gantt_chart_to_excel(self, file_path=None):
        """
        Sadece Gantt şemasını ayrı bir Excel dosyasına aktarır.

        :param file_path: Excel dosyasının kaydedileceği konum. None ise kullanıcıdan sorulur.
        :return: Başarılı olursa True, aksi halde False
        """
        try:
            # Atama çıktılarını al
            assignments = self.get_assignments_for_output()
            worker_assignments = self.get_worker_assignments()

            if not assignments:
                print("No assignments to export.")
                return False

            # Eğer dosya yolu belirtilmemişse, kullanıcıdan al
            if not file_path:
                from tkinter import filedialog
                file_path = filedialog.asksaveasfilename(
                    title="Save Gantt Chart Excel File",
                    filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")),
                    defaultextension=".xlsx"
                )

                if not file_path:  # Kullanıcı iptal ettiyse
                    return False

            # Yeni bir Excel çalışma kitabı oluştur
            wb = openpyxl.Workbook()
            # Varsayılan sayfayı sil
            wb.remove(wb.active)

            # Stil tanımlamaları
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # --- GANTT CHART SAYFASI ---
            gantt_sheet = wb.create_sheet("Worker Gantt Chart")

            # Tüm tarihleri topla ve sırala
            all_dates = set()
            for assignment in assignments:
                all_dates.add(assignment["Date"])
            sorted_dates = sorted(list(all_dates), key=lambda x: datetime.strptime(x, "%d.%m.%Y"))

            # Tüm zaman aralıklarını topla
            all_time_intervals = set()
            for assignment in assignments:
                all_time_intervals.add(assignment["Time Interval"])

            # Zaman aralıklarını saatlere göre sırala
            sorted_time_intervals = sorted(
                list(all_time_intervals),
                key=lambda x: datetime.strptime(x.split("-")[0], "%H:%M").time()
            )

            # Başlık satırı
            gantt_sheet.cell(row=1, column=1).value = "Worker Gantt Chart"
            gantt_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
            gantt_sheet.merge_cells('A1:E1')
            gantt_sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')

            # Alt başlık - İşçi bilgileri başlıkları
            gantt_sheet.cell(row=3, column=1).value = "Worker Name"
            gantt_sheet.cell(row=3, column=2).value = "Registration No"
            gantt_sheet.cell(row=3, column=3).value = "Total Hours"

            # Sütun genişliklerini ayarla
            gantt_sheet.column_dimensions['A'].width = 25  # İşçi adı sütunu genişliği
            gantt_sheet.column_dimensions['B'].width = 15  # Sicil no sütunu genişliği
            gantt_sheet.column_dimensions['C'].width = 12  # Toplam saat sütunu genişliği

            # Stil ayarlamaları - başlıklar
            for col in range(1, 4):
                cell = gantt_sheet.cell(row=3, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                cell.border = thin_border

            # Tarih ve saat başlıkları
            col_offset = 4  # İlk 3 sütun işçi bilgileri için ayrıldı
            for date_idx, date_str in enumerate(sorted_dates):
                date_col = col_offset + date_idx * (len(sorted_time_intervals) + 1)
                date_obj = datetime.strptime(date_str, "%d.%m.%Y")
                date_display = date_obj.strftime("%d.%m.%Y")

                # Tarih başlığı
                gantt_sheet.cell(row=2, column=date_col).value = date_display
                gantt_sheet.cell(row=2, column=date_col).font = Font(bold=True)
                gantt_sheet.merge_cells(
                    start_row=2,
                    start_column=date_col,
                    end_row=2,
                    end_column=date_col + len(sorted_time_intervals) - 1
                )
                gantt_sheet.cell(row=2, column=date_col).alignment = Alignment(horizontal='center')

                # Saat aralığı başlıkları
                for time_idx, time_interval in enumerate(sorted_time_intervals):
                    time_col = date_col + time_idx
                    gantt_sheet.cell(row=3, column=time_col).value = time_interval
                    gantt_sheet.cell(row=3, column=time_col).font = Font(bold=True)
                    gantt_sheet.cell(row=3, column=time_col).alignment = Alignment(horizontal='center')
                    gantt_sheet.column_dimensions[openpyxl.utils.get_column_letter(time_col)].width = 10

            # Stil ayarlamaları - tarih ve saat başlıkları
            for row in range(2, 4):
                for col in range(4, col_offset + len(sorted_dates) * (len(sorted_time_intervals) + 1)):
                    cell = gantt_sheet.cell(row=row, column=col)
                    if cell.value:  # Boş olmayan hücreler için
                        cell.border = thin_border
                        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

            # İşçileri ve atamalarını ekle
            row_offset = 4
            worker_colors = {}  # İşçi renklerini takip etmek için
            color_index = 0
            color_palette = [
                "9BC2E6",  # Açık mavi
                "A9D08E",  # Açık yeşil
                "FFD966",  # Sarı
                "F4B084",  # Turuncu
                "C9C9C9",  # Gri
                "8EA9DB",  # Mavi
                "FF99CC",  # Pembe
                "AEAAAA",  # Gümüş
            ]

            # Her işçi için toplam çalışma saatlerini hesapla
            worker_total_hours = {}
            for worker_name, data in worker_assignments.items():
                total_hours = 0

                for assignment in data["assignments"]:
                    # Zaman aralığını parçala
                    time_str = assignment["time"]
                    start_time_str, end_time_str = time_str.split("-")

                    # Saat ve dakikayı al
                    start_time = datetime.strptime(start_time_str, "%H:%M").time()
                    end_time = datetime.strptime(end_time_str, "%H:%M").time()

                    # Saat farkını hesapla (saatler ve dakikalar)
                    start_minutes = start_time.hour * 60 + start_time.minute
                    end_minutes = end_time.hour * 60 + end_time.minute

                    # Gece yarısı geçişini kontrol et
                    if end_minutes < start_minutes:  # Örneğin 23:00-01:00 durumu
                        end_minutes += 24 * 60  # Bir günlük dakikaları ekle

                    # Dakika farkını saate çevir
                    hours_diff = (end_minutes - start_minutes) / 60
                    total_hours += hours_diff

                worker_total_hours[worker_name] = total_hours

            # Tüm işçiler için sırayla satır oluştur
            for worker_name, data in worker_assignments.items():
                worker_row = row_offset
                row_offset += 1

                # İşçi adı
                gantt_sheet.cell(row=worker_row, column=1).value = worker_name
                gantt_sheet.cell(row=worker_row, column=1).alignment = Alignment(horizontal='left')
                gantt_sheet.cell(row=worker_row, column=1).border = thin_border

                # Sicil numarası
                gantt_sheet.cell(row=worker_row, column=2).value = data["registration_number"]
                gantt_sheet.cell(row=worker_row, column=2).alignment = Alignment(horizontal='center')
                gantt_sheet.cell(row=worker_row, column=2).border = thin_border

                # Toplam çalışma saati - doğru hesaplanmış hali
                total_hours = worker_total_hours[worker_name]
                gantt_sheet.cell(row=worker_row, column=3).value = round(total_hours, 2)  # 2 ondalık basamağa yuvarla
                gantt_sheet.cell(row=worker_row, column=3).alignment = Alignment(horizontal='center')
                gantt_sheet.cell(row=worker_row, column=3).border = thin_border

                # Bu işçi için renk seç
                if worker_name not in worker_colors:
                    worker_colors[worker_name] = color_palette[color_index % len(color_palette)]
                    color_index += 1

                worker_color = worker_colors[worker_name]

                # İşçinin atamalarını işle
                for assignment in data["assignments"]:
                    date_str = assignment["date"]
                    time_interval = assignment["time"]
                    operation = assignment["operation"]
                    product = assignment["product"]

                    # Tarihin indeksini bul
                    date_idx = sorted_dates.index(date_str)
                    # Zaman aralığının indeksini bul
                    time_idx = sorted_time_intervals.index(time_interval)

                    # Gantt hücresinin konumunu hesapla
                    cell_col = col_offset + date_idx * (len(sorted_time_intervals) + 1) + time_idx

                    # Hücreyi doldur
                    cell = gantt_sheet.cell(row=worker_row, column=cell_col)
                    cell.value = f"{product}-{operation}"
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color=worker_color, end_color=worker_color, fill_type="solid")
                    cell.border = thin_border

            # Açıklayıcı not ekle
            note_row = row_offset + 2
            gantt_sheet.cell(row=note_row, column=1).value = "Note: Each cell contains 'Product-Operation'"
            gantt_sheet.cell(row=note_row, column=1).font = Font(italic=True)
            gantt_sheet.merge_cells(
                start_row=note_row,
                start_column=1,
                end_row=note_row,
                end_column=5
            )

            # Boş hücrelere ince kenarlık ekle
            for row in range(4, row_offset):
                for col in range(4, col_offset + len(sorted_dates) * (len(sorted_time_intervals) + 1)):
                    cell = gantt_sheet.cell(row=row, column=col)
                    if not cell.value:  # Boş hücreler için
                        cell.border = thin_border

            # Dosyayı kaydet
            wb.save(file_path)
            print(f"Gantt chart exported to {file_path}")
            return True

        except Exception as e:
            print(f"Error exporting Gantt chart to Excel: {e}")
            import traceback
            traceback.print_exc()
            return False
    def debug(self):
        print("debug")

    def run_GUI(self):
        self.screenController = mainscreen.MainWindow()  # create GUI
        self.screenController.setMainController(self)
        print("GUI running")
        self.screenController.mainloop()


if __name__ == "__main__":
    main = MainController()
    main.run_GUI()



