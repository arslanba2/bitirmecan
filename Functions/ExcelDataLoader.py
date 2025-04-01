import openpyxl
from Models import Operation, Product
from Models import Jig
from Models.Worker import Worker
from Models.Dictionaries import SHIFT_SCHEDULES


class ExcelDataLoader:
    def __init__(self):
        self.__working_order = None  # (str) Kullanıcıdan mainscreende alınır
        self.__excel_path = None  # (file path) Kullanıcıdan mainscreende alınır
        self.__starting_shift = None  # (str) Kullanıcıdan mainscreende alınır
        self.__products = []
        self.__jigs = []
        self.__workers = []

    def set_products(self, _products):
        self.__products = _products

    def set_jigs(self, _jig):
        self.__jigs = _jig

    def set_workers(self, _worker):
        self.__workers = _worker

    def set_working_order(self, _working_order):
        self.__working_order = _working_order

    def set_starting_shift(self, _starting_shift):
        self.__starting_shift = _starting_shift

    def get_starting_shift(self):
        return self.__starting_shift

    def get_product(self, serialNumber):
        for product in self.__products:
            if product.get_serial_number() == serialNumber:
                return product

    def read_jigs_from_excel(self, _excel_path):
        self.__excel_path = _excel_path

        try:
            wb = openpyxl.load_workbook(self.__excel_path)
            sheet_operations = wb["Operasyon Bilgi"]

            for row in range(1, sheet_operations.max_row):
                jig_value = sheet_operations[f"A{row + 1}"].value
                jig_list = set() if jig_value == "-" else set(str(jig_value).split(",")) if jig_value else {
                    "No jig requirement"}
                jig_list = sorted(jig_list)
                for jig in jig_list:
                    Jig.create_jig(self.__jigs, jig)
            wb.close()
            return 1
        except Exception as e:
            return e

    def read_operations_from_excel(self, sn):
        try:
            wb = openpyxl.load_workbook(self.__excel_path)
            sheet_operations = wb["Operasyon Bilgi"]

            prdct = self.get_product(sn)
            operations_dict = {}
            for row in range(1, sheet_operations.max_row):
                op = Operation.Operation()

                jig_value = sheet_operations[f"A{row + 1}"].value
                op.set_compatible_jigs(
                    set() if jig_value == "-" else set(str(jig_value).split(",")) if jig_value else {
                        "No jig requirement"})
                op.set_name(str(sheet_operations[f"B{row + 1}"].value) if sheet_operations[
                    f"B{row + 1}"].value else "Unknown")

                op.set_required_skills(str(sheet_operations[f"C{row + 1}"].value) if sheet_operations[
                    f"C{row + 1}"].value else "None")
                op.set_required_man_hours(float(sheet_operations[f"D{row + 1}"].value or 0.0))
                op.set_min_workers(int(sheet_operations[f"E{row + 1}"].value or 1))
                op.set_max_workers(int(sheet_operations[f"F{row + 1}"].value or 1))
                predecessors = sheet_operations[f"G{row + 1}"].value
                if predecessors:
                    elements = set()
                    for part in str(predecessors).split(","):
                        if "-" in part:
                            start, end = map(int, part.split("-"))
                            elements.update(str(i) for i in range(start, end + 1))
                        else:
                            elements.add(part.strip())
                    predecessor_objects = [op for op in prdct.get_operations() if op.get_name() in elements]
                    op.set_predecessors(predecessor_objects)  # Burada Operation nesnelerini ayarlıyoruz
                    op.set_uncompleted_predecessors(predecessor_objects)
                else:
                    op.set_predecessors([])
                    op.set_uncompleted_predecessors([])

                prdct.add_operation(op)
                operations_dict[op.get_name()] = op

            for product in self.__products:
                for op_index in range(len(product.get_operations())):
                    successors_list = []
                    op = product.get_operations()[op_index]
                    for next_op_index in range(len(product.get_operations())):
                        if op.get_name() != product.get_operations()[next_op_index].get_name():  # öncül ara
                            successor_op = product.get_operations()[next_op_index]
                            for pre in successor_op.get_predecessors():
                                if pre.get_name() == op.get_name():  # Burada da Operation nesnesini kullanıyoruz
                                    successors_list.append(successor_op.get_name())
                    op.set_successors(successors_list)

            wb.close()
            return 1
        except Exception as e:
            return e

    def read_workers_from_excel(self, _excel_path):

        self.__excel_path = _excel_path

        try:
            wb = openpyxl.load_workbook(self.__excel_path)
            sheet_workers = wb["Çalışan Vardiya Matrisi"]
            sheet_fications = wb["Çalışan Yetenek Matrisi"]
            for row in range(2, sheet_workers.max_row + 1):
                w = Worker()
                w.set_registration_number(str(sheet_workers.cell(row=row, column=2).value))
                w.set_name(str(sheet_workers.cell(row=row, column=3).value))

                # Reading shift information
                shift_schedule = []
                for col in range(6, sheet_workers.max_column + 1):
                    cell_value = sheet_workers.cell(row=1, column=col).value
                    shift_type = sheet_workers.cell(row=row, column=col).value
                    if shift_type is not None:
                        # SHIFT_SCHEDULES sözlüğünden zaman aralıklarını al
                        shift_intervals = SHIFT_SCHEDULES.get(self.__working_order, {}).get(shift_type, [])

                        # Eğer zaman aralıkları bulunamazsa, bir hata mesajı göster
                        if not shift_intervals:
                            print(
                                f"Warning: No intervals found for working order '{self.__working_order}' and shift type '{shift_type}'")

                        shift_schedule.append([cell_value, shift_type, shift_intervals])

                    w.set_shift_schedule(shift_schedule)

                restriction_list = []

                for skill_row in range(2, sheet_fications.max_row + 1):
                    skill_worker_id = str(sheet_fications.cell(row=skill_row, column=2).value)
                    if skill_worker_id == w.get_registration_number():
                        skill = str(sheet_fications.cell(row=skill_row, column=4).value)
                        w.set_skills(skill)
                        restriction_value = sheet_fications.cell(row=skill_row, column=5).value
                        if restriction_value:
                            restriction_list.append(set(str(restriction_value).split(",")))
                            w.set_restrictions(restriction_list)
                        break

                self.__workers.append(w)
            wb.close()
            return 1
        except Exception as e:
            return e




